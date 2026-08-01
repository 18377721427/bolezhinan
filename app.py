# -*- coding: utf-8 -*-
"""伯乐职南 · 沙龙运营管理系统 — 主应用"""
import json, hashlib, datetime, os, socket, re
from io import BytesIO
from functools import wraps
from flask import Flask, request, jsonify, session, redirect, render_template, make_response
from models import db, now, today, Activity, ActivityVersion, Task, TaskTemplate
from models import Registration, Customer, Template, TemplateVersion, OperationLog, User, Feedback, SiteConfig
from nlp_summary import generate_summary, classify_feedback
try:
    from feishu_base import FeishuBaseClient, FeishuCliError, FeishuApiError
    HAS_FEISHU = True
except Exception:
    HAS_FEISHU = False
    FeishuBaseClient = FeishuCliError = FeishuApiError = None

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev|secret')

# 自定义 Jinja2 过滤器
@app.template_filter('fromjson')
def fromjson_filter(s):
    try: return json.loads(s)
    except: return []
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL',
    'sqlite:///' + os.path.join(os.path.dirname(os.path.abspath(__file__)), 'bolezhinan.db')
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

# ─── 辅助函数 ───
def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def wrapper(*a, **ka):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': '未登录'}), 401
            return redirect('/login')
        return f(*a, **ka)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*a, **ka):
        if 'user_id' not in session or session.get('role') not in ('admin', 'operator'):
            if request.path.startswith('/api/'):
                return jsonify({'error': '无权限'}), 403
            return redirect('/login')
        return f(*a, **ka)
    return wrapper

def log(module, action, detail=''):
    op = session.get('display_name', 'system')
    db.session.add(OperationLog(module=module, action=action, operator=op, detail=detail))
    db.session.commit()

# ─── 初始化数据库 + 种子数据 ───
def init_db():
    """初始化数据库表结构（不删除已有数据）"""
    db.create_all()
    # ── 种子数据：只在空表时填充 ──
    if not User.query.first():
        db.session.add(User(username='admin', password_hash=hash_pw('admin123'),
                            display_name='李梅', role='admin'))
        db.session.add(User(username='operator', password_hash=hash_pw('op123'),
                            display_name='搭档', role='operator'))
    if not TaskTemplate.query.first():
        phases_data = [
            ('策划期', ['确定活动主题', '邀请嘉宾', '确定场地', '制定预算']),
            ('准备期', ['海报设计', '邀约话术', '签到表设计', '物料采购', 'PPT 准备']),
            ('推广期', ['朋友圈宣传', '社群推送', '小红书发帖', '一对一邀约', '粗门发布']),
            ('执行期', ['现场布置', '签到引导', '主持串场', '摄影记录']),
            ('收尾期', ['活动群维护', '照片整理', '感谢信发送', '物资归位']),
            ('复盘期', ['数据汇总', '团队复盘会', '反馈收集', '下期规划']),
        ]
        order = 0
        for phase, tasks in phases_data:
            for tname in tasks:
                db.session.add(TaskTemplate(name=tname, phase=phase, sort_order=order))
                order += 1
    if Template.query.count() == 0:
        # 不再自动生成示例模板
        pass
    if Activity.query.count() == 0:
        # 不再自动生成示例活动
        pass
    if not SiteConfig.query.filter_by(key='salon_intro').first():
        db.session.add(SiteConfig(key='salon_intro',
            value='伯乐职南职场沙龙——为职场人打造的线下交流平台，每期围绕一个职场话题深度探讨，链接优质人脉，助力职业发展。'))
    
# ─── 认证路由 ───
@app.route('/login')
def login_page():
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json()
    user = User.query.filter_by(username=data.get('username','')).first()
    if user and user.password_hash == hash_pw(data.get('password','')):
        session['user_id'] = user.id
        session['role'] = user.role
        session['display_name'] = user.display_name
        return jsonify({'ok': True, 'user': user.to_dict()})
    return jsonify({'error': '用户名或密码错误'}), 401

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

def get_lan_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

@app.route('/admin')
@login_required
def admin_page():
    resp = make_response(render_template('admin.html', display_name=session.get('display_name',''),
                           role=session.get('role',''), now=datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
                           lan_ip=get_lan_ip()))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

@app.route('/register/<int:activity_id>')
def register_page(activity_id):
    act = Activity.query.get(activity_id)
    if not act:
        return '活动不存在', 404
    return render_template('register.html', act=act)

# ─── API: Dashboard ───
@app.route('/api/dashboard/stats')
@login_required
def api_dashboard_stats():
    total_activities = Activity.query.filter(Activity.deleted_at.is_(None)).count()
    total_regs = Registration.query.count()
    checked = Registration.query.filter_by(checkin_status='checked').count()
    attendance = round(checked / total_regs * 100) if total_regs else 0
    repeat_customers = Customer.query.filter(Customer.total_participations >= 2).count()
    total_customers = Customer.query.count()
    repeat_rate = round(repeat_customers / total_customers * 100) if total_customers else 0
    this_month_start = datetime.date.today().strftime('%Y-%m') + '-01'
    new_customers = Customer.query.filter(Customer.created_at >= this_month_start).count()
    # 待办事项
    todos = Task.query.filter(Task.status.in_(['todo','in_progress'])).order_by(Task.due_date).limit(5).all()
    overdue = Task.query.filter(Task.due_date != '', Task.due_date.isnot(None),
                                  Task.due_date < today(), Task.status != 'done').count()
    # 活动提醒：明天有活动吗？
    import datetime as dt
    tomorrow = (dt.date.today() + dt.timedelta(days=1)).strftime('%Y-%m-%d')
    upcoming_acts = Activity.query.filter(Activity.date == tomorrow, Activity.deleted_at.is_(None),
                                           Activity.status.in_(['published','ongoing'])).all()
    reminders = []
    for act in upcoming_acts:
        regs = Registration.query.filter_by(activity_id=act.id)
        total_regs = regs.count()
        remind_list = regs.filter(Registration.wechat != '').all()
        reminders.append({
            'activity_id': act.id,
            'name': act.name,
            'date': act.date,
            'time': act.time,
            'venue': act.venue,
            'registrations': total_regs,
            'to_remind': len(remind_list),
        })
    return jsonify({
        'total_activities': total_activities,
        'total_registrations': total_regs,
        'attendance_rate': attendance,
        'repeat_rate': repeat_rate,
        'new_customers': new_customers,
        'total_customers': total_customers,
        'checked_in': checked,
        'overdue_tasks': overdue,
        'todos': [t.to_dict() for t in todos],
        'reminders': reminders,
    })

@app.route('/api/dashboard/trends')
@login_required
def api_dashboard_trends():
    acts = Activity.query.filter(Activity.deleted_at.is_(None), Activity.status == 'ended').order_by(Activity.date).all()
    trends = []
    for a in acts:
        regs = Registration.query.filter_by(activity_id=a.id)
        total = regs.count()
        ckd = regs.filter_by(checkin_status='checked').count()
        rate = round(ckd / total * 100) if total else 0
        trends.append({'name': a.name[:10]+'…' if len(a.name) > 10 else a.name, 'rate': rate, 'total': total})
    return jsonify(trends)

@app.route('/api/dashboard/channels')
@login_required
def api_dashboard_channels():
    channels = {}
    for r in Registration.query.all():
        ch = r.source_channel or '其他'
        channels[ch] = channels.get(ch, 0) + 1
    total = sum(channels.values()) or 1
    result = [{'name': k, 'count': v, 'percent': round(v/total*100)} for k, v in channels.items()]
    return jsonify(result)

# ─── API: 活动 ───
@app.route('/api/activities')
@login_required
def api_activities():
    status_f = request.args.get('status', '')
    q = Activity.query.filter(Activity.deleted_at.is_(None))
    if status_f:
        q = q.filter_by(status=status_f)
    acts = q.order_by(Activity.created_at.desc()).all()
    result = []
    for a in acts:
        d = a.to_dict()
        regs = Registration.query.filter_by(activity_id=a.id)
        d['registration_count'] = regs.count()
        d['checked_count'] = regs.filter_by(checkin_status='checked').count()
        d['attendance'] = round(d['checked_count'] / d['registration_count'] * 100) if d['registration_count'] else 0
        result.append(d)
    return jsonify(result)

@app.route('/api/activities', methods=['POST'])
@login_required
def api_create_activity():
    data = request.get_json()
    act = Activity(
        name=data.get('name','新建活动'), theme=data.get('theme',''),
        date=data.get('date',''), time=data.get('time',''), venue=data.get('venue',''),
        guest_desc=data.get('guest_desc',''), ticket_price=float(data.get('ticket_price',0)),
        max_participants=int(data.get('max_participants',30)),
        status='draft', based_on_id=data.get('based_on_id'),
        desc_text=data.get('desc_text',''), highlight_text=data.get('highlight_text',''),
        tags=data.get('tags',''))
    db.session.add(act)
    db.session.flush()
    # 复制任务
    tmpls = TaskTemplate.query.order_by(TaskTemplate.sort_order).all()
    for tmpl in tmpls:
        db.session.add(Task(activity_id=act.id, name=tmpl.name, phase=tmpl.phase,
                            status='todo', sort_order=tmpl.sort_order))
    log('活动', f'创建活动 #{act.id} {act.name}')
    db.session.commit()
    return jsonify(act.to_dict())

@app.route('/api/activities/<int:aid>')
@login_required
def api_activity_detail(aid):
    act = Activity.query.get_or_404(aid)
    return jsonify(act.to_dict())

@app.route('/api/activities/<int:aid>', methods=['PUT'])
@login_required
def api_update_activity(aid):
    act = Activity.query.get_or_404(aid)
    for f in ('name','theme','date','time','venue','guest_desc','status',
              'desc_text','highlight_text','tags','flow_json','max_participants'):
        if f in request.json:
            setattr(act, f, request.json[f])
    if 'ticket_price' in request.json:
        act.ticket_price = float(request.json['ticket_price'])
    act.updated_at = now()
    log('活动', f'编辑活动 #{act.id} {act.name}')
    db.session.commit()
    return jsonify(act.to_dict())

@app.route('/api/activities/<int:aid>', methods=['DELETE'])
@login_required
def api_delete_activity(aid):
    act = Activity.query.get_or_404(aid)
    act.deleted_at = now()
    log('活动', f'删除活动 #{act.id} {act.name}')
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/activities/<int:aid>/publish', methods=['POST'])
@login_required
def api_publish_activity(aid):
    act = Activity.query.get_or_404(aid)
    act.status = 'published'
    act.version = (act.version or 0) + 1
    # 保存版本快照
    db.session.add(ActivityVersion(activity_id=aid, snapshot_json=json.dumps(act.to_dict(), ensure_ascii=False),
                                    version=act.version, created_by=session.get('display_name','admin')))
    log('活动', f'发布活动 #{act.id} {act.name} v{act.version}')
    db.session.commit()
    return jsonify(act.to_dict())

@app.route('/api/activities/<int:aid>/versions')
@login_required
def api_activity_versions(aid):
    vs = ActivityVersion.query.filter_by(activity_id=aid).order_by(ActivityVersion.version.desc()).all()
    return jsonify([{'id': v.id, 'version': v.version, 'created_at': v.created_at,
                     'created_by': v.created_by} for v in vs])

@app.route('/api/activities/<int:aid>/rollback', methods=['POST'])
@login_required
def api_rollback_activity(aid):
    ver = request.json.get('version')
    v = ActivityVersion.query.filter_by(activity_id=aid, version=ver).first()
    if not v:
        return jsonify({'error': '版本不存在'}), 404
    snap = json.loads(v.snapshot_json)
    act = Activity.query.get(aid)
    for k, val in snap.items():
        if k not in ('id', 'created_at', 'deleted_at'):
            setattr(act, k, val)
    act.version = (act.version or 0) + 1
    log('活动', f'回退活动 #{aid} 到 v{ver}')
    db.session.commit()
    return jsonify(act.to_dict())

# ─── API: 任务 ───
@app.route('/api/activities/<int:aid>/tasks')
@login_required
def api_tasks(aid):
    phase = request.args.get('phase', '')
    q = Task.query.filter_by(activity_id=aid)
    if phase:
        q = q.filter_by(phase=phase)
    ts = q.order_by(Task.sort_order).all()
    # 逾期检测
    td = today()
    for t in ts:
        if t.due_date and t.due_date.strip() and t.due_date < td and t.status == 'todo':
            t.status = 'overdue'
    db.session.commit()
    return jsonify([t.to_dict() for t in ts])

@app.route('/api/tasks/<int:tid>', methods=['PUT'])
@login_required
def api_update_task(tid):
    t = Task.query.get_or_404(tid)
    for f in ('status', 'assignee', 'due_date', 'name'):
        if f in request.json:
            setattr(t, f, request.json[f])
    t.updated_at = now()
    log('任务', f'更新任务 #{tid} → {request.json}')
    db.session.commit()
    return jsonify(t.to_dict())

@app.route('/api/tasks', methods=['POST'])
@login_required
def api_create_task():
    d = request.json
    max_order = db.session.query(db.func.max(Task.sort_order)).filter_by(activity_id=d['activity_id']).scalar() or 0
    t = Task(activity_id=d['activity_id'], name=d['name'], phase=d.get('phase','策划期'),
             assignee=d.get('assignee',''), due_date=d.get('due_date',''),
             status='todo', sort_order=max_order + 1)
    db.session.add(t)
    log('任务', f'添加自定义任务 {d["name"]}')
    db.session.commit()
    return jsonify(t.to_dict())

# ─── API: 报名 ───
@app.route('/api/registrations')
@login_required
def api_registrations():
    aid = request.args.get('activity_id', type=int)
    search = request.args.get('search', '')
    q = Registration.query
    if aid:
        q = q.filter_by(activity_id=aid)
    if search:
        q = q.filter(db.or_(Registration.name.contains(search), Registration.phone.contains(search), Registration.wechat.contains(search)))
    regs = q.order_by(Registration.created_at.desc()).all()
    return jsonify([r.to_dict() for r in regs])

@app.route('/api/registrations/<int:rid>', methods=['PUT'])
@login_required
def api_update_registration(rid):
    r = Registration.query.get_or_404(rid)
    if 'payment_status' in request.json:
        r.payment_status = request.json['payment_status']
    if 'checkin_status' in request.json:
        r.checkin_status = request.json['checkin_status']
    if 'name' in request.json:
        r.name = request.json['name']
    if 'phone' in request.json:
        r.phone = request.json['phone']
    if 'wechat' in request.json:
        r.wechat = request.json['wechat']
    if 'source_channel' in request.json:
        r.source_channel = request.json['source_channel']
    log('报名', f'更新报名 #{rid} → {request.json}')
    db.session.commit()
    return jsonify(r.to_dict())

@app.route('/api/registrations/<int:rid>', methods=['DELETE'])
@login_required
def api_delete_registration(rid):
    r = Registration.query.get_or_404(rid)
    db.session.delete(r)
    log('报名', f'删除报名 #{rid}')
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/register', methods=['POST'])
def api_register():
    d = request.json
    if not d.get('name') or not d.get('phone'):
        return jsonify({'error': '姓名和手机号/微信号为必填项'}), 400
    phone = d['phone']
    wechat = d.get('wechat', '') or phone
    # 如果输入的是纯11位数字则视为手机号，否则视为微信号
    if phone.isdigit() and len(phone) == 11:
        pass  # valid phone
    else:
        # 非手机号格式，把phone当微信号，留空phone
        wechat = phone
        phone = ''
    act = Activity.query.get(d.get('activity_id'))
    if not act:
        return jsonify({'error': '活动不存在'}), 404
    # 检查名额
    reg_count = Registration.query.filter_by(activity_id=act.id).count()
    if act.max_participants and reg_count >= act.max_participants:
        return jsonify({'error': '活动名额已满'}), 400
    # 查找/创建客户
    wechat = d['wechat']
    customer = Customer.query.filter_by(phone=phone).first()
    if customer:
        customer.total_participations += 1
        customer.wechat = wechat or customer.wechat
        customer.updated_at = now()
    else:
        customer = Customer(name=d['name'], phone=phone, wechat=wechat,
                            source_channel=d.get('source_channel',''), total_participations=1,
                            first_activity_id=act.id)
        db.session.add(customer)
        db.session.flush()
    # 创建报名（默认 unpaid）
    reg = Registration(activity_id=act.id, customer_id=customer.id,
                       name=d['name'], phone=phone, wechat=wechat,
                       source_channel=d.get('source_channel',''),
                       payment_status='unpaid')
    db.session.add(reg)
    db.session.commit()
    return jsonify({'ok': True, 'name': d['name'], 'activity_name': act.name, 'id': reg.id})

# ─── API: 客户 ───
@app.route('/api/customers')
@login_required
def api_customers():
    search = request.args.get('search', '')
    min_part = request.args.get('min_part', type=int)
    topic = request.args.get('topic', '')
    q = Customer.query
    if search:
        q = q.filter(db.or_(Customer.name.contains(search), Customer.phone.contains(search), Customer.wechat.contains(search)))
    if min_part:
        q = q.filter(Customer.total_participations >= min_part)
    if topic:
        # Find activity ids with this theme
        act_ids = [a.id for a in Activity.query.filter(Activity.theme.contains(topic)).all()]
        if act_ids:
            # Find customer ids who registered for these activities
            cust_ids = db.session.query(Registration.customer_id).filter(
                Registration.activity_id.in_(act_ids),
                Registration.customer_id.isnot(None)
            ).distinct().all()
            cust_ids = [c[0] for c in cust_ids if c[0]]
            if cust_ids:
                q = q.filter(Customer.id.in_(cust_ids))
            else:
                return jsonify([])
    cs = q.order_by(Customer.created_at.desc()).all()
    result = []
    for c in cs:
        d = c.to_dict()
        if d['first_activity_id']:
            fa = Activity.query.get(d['first_activity_id'])
            d['first_activity'] = fa.name if fa else ''
        else:
            d['first_activity'] = ''
        result.append(d)
    return jsonify(result)

@app.route('/api/customers/<int:cid>')
@login_required
def api_customer_detail(cid):
    c = Customer.query.get_or_404(cid)
    regs = Registration.query.filter_by(customer_id=cid).order_by(Registration.created_at.desc()).all()
    d = c.to_dict()
    d['registrations'] = [r.to_dict() for r in regs]
    return jsonify(d)

@app.route('/api/customers', methods=['POST'])
@login_required
def api_create_customer():
    d = request.json
    c = Customer(name=d['name'], phone=d.get('phone',''), wechat=d.get('wechat',''),
                 source_channel=d.get('source_channel',''))
    db.session.add(c)
    log('客户', f'手动添加客户 {d["name"]}')
    db.session.commit()
    return jsonify(c.to_dict())

@app.route('/api/customers/<int:cid>', methods=['PUT'])
@login_required
def api_update_customer(cid):
    c = Customer.query.get_or_404(cid)
    for f in ('name', 'phone', 'wechat', 'source_channel'):
        if f in request.json:
            setattr(c, f, request.json[f])
    c.updated_at = now()
    log('客户', f'编辑客户 #{cid} → {request.json}')
    db.session.commit()
    return jsonify(c.to_dict())

@app.route('/api/customers/<int:cid>', methods=['DELETE'])
@login_required
def api_delete_customer(cid):
    c = Customer.query.get_or_404(cid)
    db.session.delete(c)
    log('客户', f'删除客户 #{cid}')
    db.session.commit()
    return jsonify({'ok': True})

# ─── API: 模板 ───
@app.route('/api/templates')
@login_required
def api_templates():
    ts = Template.query.all()
    return jsonify([t.to_dict() for t in ts])

@app.route('/api/templates/<int:tid>', methods=['PUT'])
@login_required
def api_update_template(tid):
    t = Template.query.get_or_404(tid)
    t.content = request.json.get('content', t.content)
    t.current_version += 1
    db.session.add(TemplateVersion(template_id=tid, content_snapshot=t.content,
                                    version=t.current_version,
                                    created_by=session.get('display_name','admin')))
    t.updated_at = now()
    log('模板', f'编辑模板 #{tid} v{t.current_version}')
    db.session.commit()
    return jsonify(t.to_dict())

@app.route('/api/templates', methods=['POST'])
@login_required
def api_create_template():
    d = request.json
    t = Template(name=d['name'], type=d.get('type',''), content=d.get('content',''))
    db.session.add(t)
    log('模板', f'创建模板 {d["name"]}')
    db.session.commit()
    return jsonify(t.to_dict())

# ─── API: 操作日志 ───
@app.route('/api/logs')
@login_required
def api_logs():
    logs = OperationLog.query.order_by(OperationLog.created_at.desc()).limit(50).all()
    return jsonify([l.to_dict() for l in logs])

# ─── API: 用户信息 ───
@app.route('/api/me')
@login_required
def api_me():
    user = User.query.get(session['user_id'])
    return jsonify(user.to_dict() if user else {})

# ─── API: Excel模板下载 ───
def _make_excel(wb, filename):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/api/templates/excel/checkin')
def api_excel_checkin():
    import openpyxl
    from flask import send_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '签到表'
    ws.merge_cells('A1:F1')
    ws['A1'] = '活动签到表'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    ws['A2'] = '活动名称：______________    日期：______________    地点：______________'
    ws.merge_cells('A2:F2')
    headers = ['编号', '姓名', '手机号', '签到时间', '付款状态', '备注']
    ws.append([])
    ws.append(headers)
    for c in range(1, 7):
        cell = ws.cell(row=4, column=c)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='C9A961', end_color='C9A961', fill_type='solid')
    for i in range(1, 16):
        ws.append([i, '', '', '', '未付款', ''])
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    return send_file(_make_excel(wb, '签到表模板.xlsx'),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='签到表模板.xlsx')

@app.route('/api/templates/excel/flow')
def api_excel_flow():
    import openpyxl
    from flask import send_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '活动流程'
    ws.merge_cells('A1:D1')
    ws['A1'] = '活动流程表'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    ws['A2'] = '活动名称：______________    日期：______________'
    ws.merge_cells('A2:D2')
    headers = ['时间', '环节名称', '内容说明', '负责人/备注']
    ws.append([])
    ws.append(headers)
    for c in range(1, 5):
        cell = ws.cell(row=4, column=c)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='C9A961', end_color='C9A961', fill_type='solid')
    examples = [
        ['14:00-14:30', '签到入场·破冰互动', '到场签到，领取姓名贴，参与破冰小游戏', ''],
        ['14:30-15:30', '嘉宾分享·转型故事', '三位嘉宾分别分享自己的转型经历', ''],
        ['15:30-16:20', '圆桌讨论·Q&A', '与嘉宾面对面交流，提出你的困惑', ''],
        ['16:20-16:50', '自由交流·资源链接', '自由社交时间，结识同行伙伴', ''],
        ['16:50-17:00', '总结·合影', '活动总结，全员合影留念', ''],
    ]
    for row in examples:
        ws.append(row)
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 16
    return send_file(_make_excel(wb, '活动流程模板.xlsx'),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='活动流程模板.xlsx')

@app.route('/api/templates/excel/materials')
def api_excel_materials():
    import openpyxl
    from flask import send_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '物料清单'
    ws.merge_cells('A1:E1')
    ws['A1'] = '活动物料清单'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    ws['A2'] = '活动名称：______________    日期：______________'
    ws.merge_cells('A2:E2')
    headers = ['编号', '物料名称', '数量', '单价(元)', '备注']
    ws.append([])
    ws.append(headers)
    for c in range(1, 6):
        cell = ws.cell(row=4, column=c)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='C9A961', end_color='C9A961', fill_type='solid')
    examples = [
        ['1', '姓名贴', '30', '', ''],
        ['2', '签到笔', '5', '', ''],
        ['3', '茶歇饮品', '30', '', ''],
        ['4', '活动流程单', '30', '', '打印'],
        ['5', '纪念礼品', '30', '', ''],
    ]
    for row in examples:
        ws.append(row)
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    return send_file(_make_excel(wb, '物料清单模板.xlsx'),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='物料清单模板.xlsx')

# ─── API: 反馈分析 ───
@app.route('/api/feedbacks', methods=['GET'])
@login_required
def api_feedbacks():
    aid = request.args.get('activity_id', type=int)
    q = Feedback.query
    if aid:
        q = q.filter_by(activity_id=aid)
    fs = q.order_by(Feedback.created_at.desc()).all()
    return jsonify([f.to_dict() for f in fs])

@app.route('/api/feedbacks', methods=['POST'])
@login_required
def api_add_feedback():
    d = request.json
    if not d.get('raw_text'):
        return jsonify({'error': '请输入反馈内容'}), 400
    f = Feedback(activity_id=d.get('activity_id'), customer_name=d.get('customer_name',''),
                 raw_text=d['raw_text'], category=d.get('category',''))
    db.session.add(f)
    log('反馈', f'添加反馈: {d["raw_text"][:30]}')
    db.session.commit()
    return jsonify(f.to_dict())

@app.route('/api/feedbacks/<int:fid>', methods=['DELETE'])
@login_required
def api_delete_feedback(fid):
    f = Feedback.query.get_or_404(fid)
    db.session.delete(f)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/feedbacks/summary')
@login_required
def api_feedback_summary():
    aid = request.args.get('activity_id', type=int)
    q = Feedback.query
    if aid:
        q = q.filter_by(activity_id=aid)
    all_fb = q.all()
    if not all_fb:
        return jsonify({'pros': [], 'cons': [], 'wants': [], 'count': 0,
                        'pros_text': '暂无数据', 'cons_text': '暂无数据', 'wants_text': '暂无数据',
                        'pros_summary': '', 'cons_summary': '', 'wants_summary': '',
                        'ai_summary': '暂无数据', 'keywords': [], 'trends': []})

    result = generate_summary(all_fb)
    result['count'] = len(all_fb)
    return jsonify(result)

# ─── API: 站点配置 ───
@app.route('/api/config/<key>')
def api_get_config(key):
    c = SiteConfig.query.filter_by(key=key).first()
    return jsonify({'key': key, 'value': c.value if c else ''})

@app.route('/api/config/<key>', methods=['PUT'])
@login_required
def api_set_config(key):
    v = request.json.get('value', '')
    c = SiteConfig.query.filter_by(key=key).first()
    if c:
        c.value = v
        c.updated_at = now()
    else:
        db.session.add(SiteConfig(key=key, value=v))
    log('配置', f'更新 {key}')
    db.session.commit()
    return jsonify({'ok': True})

# ─── API: 报名导出Excel ───
@app.route('/api/registrations/export/<int:aid>')
@login_required
def api_export_registrations(aid):
    import openpyxl
    from flask import send_file
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '报名表'
    act = Activity.query.get(aid)
    ws.merge_cells('A1:G1')
    ws['A1'] = f'报名表 - {act.name if act else "活动"}'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    headers = ['姓名', '手机号', '微信号', '来源渠道', '报名时间', '付款状态', '签到状态']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='C9A961', end_color='C9A961', fill_type='solid')
    regs = Registration.query.filter_by(activity_id=aid).order_by(Registration.created_at).all()
    for i, r in enumerate(regs):
        ws.cell(row=4+i, column=1, value=r.name)
        ws.cell(row=4+i, column=2, value=r.phone)
        ws.cell(row=4+i, column=3, value=r.wechat or '')
        ws.cell(row=4+i, column=4, value=r.source_channel or '')
        ws.cell(row=4+i, column=5, value=r.created_at or '')
        ws.cell(row=4+i, column=6, value='已付款' if r.payment_status=='paid' else '未付款')
        ws.cell(row=4+i, column=7, value='已签到' if r.checkin_status=='checked' else '未签到')
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'报名表_{act.name if act else aid}.xlsx')

# ─── API: 签到表导出Excel ───
@app.route('/api/checkin/export/<int:aid>')
@login_required
def api_export_checkin(aid):
    import openpyxl
    from flask import send_file
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '签到表'
    act = Activity.query.get(aid)
    ws.merge_cells('A1:F1')
    ws['A1'] = f'签到表 - {act.name if act else "活动"}'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    headers = ['姓名', '手机号', '微信号', '来源渠道', '签到状态', '付款状态']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='C9A961', end_color='C9A961', fill_type='solid')
    regs = Registration.query.filter_by(activity_id=aid).order_by(Registration.created_at).all()
    for i, r in enumerate(regs):
        ws.cell(row=4+i, column=1, value=r.name)
        ws.cell(row=4+i, column=2, value=r.phone)
        ws.cell(row=4+i, column=3, value=r.wechat or '')
        ws.cell(row=4+i, column=4, value=r.source_channel or '')
        ws.cell(row=4+i, column=5, value='已签到' if r.checkin_status=='checked' else '未签到')
        ws.cell(row=4+i, column=6, value='已付款' if r.payment_status=='paid' else '未付款')
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'签到表_{act.name if act else aid}.xlsx')

# ─── API: 反馈导出Excel ───
@app.route('/api/feedbacks/export')
@login_required
def api_export_feedbacks():
    import openpyxl
    from flask import send_file
    from io import BytesIO
    aid = request.args.get('activity_id', type=int)
    q = Feedback.query
    if aid:
        q = q.filter_by(activity_id=aid)
    all_fb = q.order_by(Feedback.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '反馈总结'
    ws.merge_cells('A1:E1')
    ws['A1'] = '反馈总结报告'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)

    # Summary section
    pros = [f for f in all_fb if f.category == '优点' or (not f.category and any(k in f.raw_text for k in ['好','赞','喜欢','棒','收获','不错','满意','有用','感谢']))]
    cons = [f for f in all_fb if f.category == '待提高' or (not f.category and any(k in f.raw_text for k in ['不好','差','建议','改进','希望','可以更','不足','缺']))]
    wants = [f for f in all_fb if f not in pros and f not in cons]

    ws['A3'] = '👍 优点总结'
    ws['A3'].font = openpyxl.styles.Font(bold=True, size=12, color='3aaf7a')
    ws['A4'] = '；'.join([f.raw_text for f in pros]) if pros else '暂无数据'
    ws.merge_cells('A4:E4')

    ws['A6'] = '📈 待提高总结'
    ws['A6'].font = openpyxl.styles.Font(bold=True, size=12, color='e8944a')
    ws['A7'] = '；'.join([f.raw_text for f in cons]) if cons else '暂无数据'
    ws.merge_cells('A7:E7')

    ws['A9'] = '🔮 想了解的内容总结'
    ws['A9'].font = openpyxl.styles.Font(bold=True, size=12, color='4a8fd4')
    ws['A10'] = '；'.join([f.raw_text for f in wants]) if wants else '暂无数据'
    ws.merge_cells('A10:E10')

    # Detail table
    ws['A12'] = '反馈明细'
    ws['A12'].font = openpyxl.styles.Font(bold=True, size=12)
    headers = ['序号', '内容', '分类', '活动ID', '时间']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=13, column=c, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='C9A961', end_color='C9A961', fill_type='solid')
    for i, f in enumerate(all_fb):
        ws.cell(row=14+i, column=1, value=i+1)
        ws.cell(row=14+i, column=2, value=f.raw_text)
        ws.cell(row=14+i, column=3, value=f.category or '未分类')
        ws.cell(row=14+i, column=4, value=f.activity_id or '')
        ws.cell(row=14+i, column=5, value=f.created_at or '')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='反馈总结报告.xlsx')

# ─── API: 上传文件解析反馈 ───
@app.route('/api/feedbacks/upload', methods=['POST'])
@login_required
def api_upload_feedback_file():
    if 'file' not in request.files:
        return jsonify({'error': '请选择文件'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': '请选择文件'}), 400

    texts = []
    filename = f.filename.lower()

    try:
        if filename.endswith('.txt'):
            content = f.read().decode('utf-8', errors='ignore')
            texts = [t.strip() for t in content.split('\n') if t.strip()]

        elif filename.endswith('.docx'):
            import zipfile
            with zipfile.ZipFile(f) as z:
                xml = z.read('word/document.xml').decode('utf-8', errors='ignore')
                texts = re.findall(r'<w:t[^>]*>([^<]+)</w:t>', xml)
            texts = [t.strip() for t in texts if t.strip()]

        elif filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and str(cell).strip():
                        texts.append(str(cell).strip())

        elif filename.endswith('.pdf'):
            try:
                import PyPDF2
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    t = page.extract_text()
                    if t:
                        texts.extend([l.strip() for l in t.split('\n') if l.strip()])
            except:
                try:
                    import pdfplumber
                    with pdfplumber.open(f) as pdf:
                        for page in pdf.pages:
                            t = page.extract_text()
                            if t:
                                texts.extend([l.strip() for l in t.split('\n') if l.strip()])
                except:
                    return jsonify({'error': 'PDF解析失败，请尝试转换为Word或TXT格式后重新上传'}), 400
        else:
            return jsonify({'error': '不支持的文件格式，请上传Word/Excel/PDF/TXT'}), 400

        if not texts:
            return jsonify({'error': '未能从文件中读取到文本内容'}), 400

        # ── 智能整理归纳：去重、过滤噪音、自动分类 ──

        # 1. 清理：移除过短内容（<4字）、纯数字、纯标点
        cleaned = []
        for t in texts:
            t = t.strip()
            if len(t) < 4: continue
            if re.match(r'^[\d\s\-—_,，。.、/\\()（）]+$', t): continue
            # 跳过页眉页脚等常见噪音
            if re.match(r'^第\d+页|^\d+/\d+|^www\.|^http', t, re.I): continue
            cleaned.append(t)

        # 2. 去重：相似度高的只保留一条
        def similarity(a, b):
            """简易Jaccard相似度"""
            set_a, set_b = set(a), set(b)
            if not set_a or not set_b: return 0
            return len(set_a & set_b) / len(set_a | set_b)

        deduped = []
        for t in cleaned:
            is_dup = False
            for existing in deduped:
                if similarity(t, existing) > 0.7:
                    is_dup = True
                    break
            if not is_dup:
                deduped.append(t)

        # 3. 自动分类关键词映射
        category_keywords = {
            '优点': ['好', '棒', '赞', '喜欢', '满意', '收获', '受益', '精彩', '有趣', '实用',
                     '干货', '不错', '开心', '感谢', '学到', '启发', '提升', '感动', '专业', 'nice', 'good'],
            '待提高': ['建议', '希望', '不足', '改进', '太短', '太长', '不够', '可惜', '遗憾',
                      '混乱', '嘈杂', '改善', '调整', '增加', '减少', '不好', '一般', '无聊', '差'],
            '想了解': ['想了解', '想知道', '感兴趣', '下次', '期待', '希望下次', '有没有',
                      '能否', '是否', '什么', '如何', '怎么', '请教', '问', '主题', '话题']
        }

        def auto_category(text):
            text_lower = text.lower()
            for cat, keywords in category_keywords.items():
                for kw in keywords:
                    if kw in text_lower:
                        return cat
            return ''

        # 4. 合并短文本为完整句子
        merged = []
        buffer = ''
        for t in deduped:
            # 如果以句号/问号/感叹号/换行结尾，或是独立长句
            if any(t.endswith(p) for p in '。！？.!?\n') or len(t) > 20:
                if buffer:
                    merged.append(buffer + t)
                    buffer = ''
                else:
                    merged.append(t)
            else:
                if buffer:
                    buffer += '，' + t
                else:
                    buffer = t
        if buffer:
            merged.append(buffer)

        # 5. 保存
        count = 0
        for t in merged:
            if len(t) < 4: continue
            cat = classify_feedback(t)
            fb = Feedback(raw_text=t, category=cat,
                          activity_id=request.form.get('activity_id', type=int))
            db.session.add(fb)
            count += 1
        db.session.commit()
        log('反馈', f'上传文件导入 {count} 条反馈（原始{len(texts)}条 -> 去重整理{len(merged)}条）')
        return jsonify({'ok': True, 'count': count,
                        'preview': merged[:5],
                        'summary': f'共处理{len(texts)}条，去重整理为{len(merged)}条，自动分类{count}条'})

    except Exception as e:
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 400

# ─── API: QR码上传 ───
QRCODE_PATH = os.path.join(os.path.dirname(__file__), 'static', 'qrcode.png')
MASCOT_PATH = os.path.join(os.path.dirname(__file__), 'static', 'mascot.png')

@app.route('/api/mascot/upload', methods=['POST', 'DELETE'])
@login_required
def api_upload_mascot():
    if request.method == 'DELETE':
        if os.path.exists(MASCOT_PATH):
            os.remove(MASCOT_PATH)
            log('系统', '删除吉祥物')
        return jsonify({'ok': True})
    if 'qrcode' not in request.files:
        return jsonify({'error': '请选择图片'}), 400
    f = request.files['qrcode']
    if f.filename == '':
        return jsonify({'error': '请选择图片'}), 400
    f.save(MASCOT_PATH)
    log('系统', '上传吉祥物')
    return jsonify({'ok': True})

@app.route('/api/mascot/check')
def api_check_mascot():
    return jsonify({'exists': os.path.exists(MASCOT_PATH)})

@app.route('/api/qrcode', methods=['GET', 'DELETE'])
@login_required
def api_get_qrcode():
    if request.method == 'DELETE':
        if os.path.exists(QRCODE_PATH):
            os.remove(QRCODE_PATH)
            log('系统', '删除微信二维码')
        return jsonify({'ok': True})
    if os.path.exists(QRCODE_PATH):
        return '', 200
    return '', 404

@app.route('/api/qrcode', methods=['DELETE'])
@login_required
def api_delete_qrcode():
    if os.path.exists(QRCODE_PATH):
        os.remove(QRCODE_PATH)
        log('系统', '删除微信二维码')
    return jsonify({'ok': True})

@app.route('/api/qrcode', methods=['POST'])
@login_required
def api_upload_qrcode():
    if 'qrcode' not in request.files:
        return jsonify({'error': '请选择图片'}), 400
    f = request.files['qrcode']
    if f.filename == '':
        return jsonify({'error': '请选择图片'}), 400
    f.save(QRCODE_PATH)
    log('系统', '上传微信二维码')
    return jsonify({'ok': True})

@app.route('/api/qrcode/check')
def api_check_qrcode():
    return jsonify({'exists': os.path.exists(QRCODE_PATH)})

# ─── API: 往期沙龙数据Excel下载 ───
@app.route('/api/templates/excel/salons')
@login_required
def api_excel_salons():
    from flask import send_file
    excel_path = os.path.join(os.path.dirname(__file__), 'static', '伯乐职南_往期沙龙数据.xlsx')
    if not os.path.exists(excel_path):
        return jsonify({'error': '文件未生成'}), 404
    return send_file(excel_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='伯乐职南_往期沙龙数据.xlsx')

# ─── API: 往期物料Excel下载 ───
@app.route('/api/templates/excel/past/<name>')
def api_excel_past(name):
    from flask import send_file
    filename_map = {
        'checkin1': '往期签到表_第一期.xlsx',
        'checkin2': '往期签到表_第二期.xlsx',
        'flow1': '往期活动流程_第一期.xlsx',
        'flow2': '往期活动流程_第二期.xlsx',
    }
    fn = filename_map.get(name)
    if not fn:
        return jsonify({'error': '未知文件'}), 404
    path = os.path.join(os.path.dirname(__file__), 'static', fn)
    if not os.path.exists(path):
        return jsonify({'error': '文件不存在'}), 404
    return send_file(path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fn)

# ─── API: 上传Excel并解析 ───
@app.route('/api/templates/upload-excel', methods=['POST'])
@login_required
def api_upload_excel():
    if 'file' not in request.files:
        return jsonify({'error': '请选择文件'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': '请选择文件'}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue  # skip header
            vals = [str(v).strip() if v else '' for v in row]
            if any(v for v in vals):
                rows.append(vals)
        return jsonify({'ok': True, 'rows': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 400

# ─── API: 飞书多维表格 ───

@app.route('/api/feishu/status')
@login_required
def api_feishu_status():
    """检查飞书CLI可用性、配置状态、已发现的表格"""
    if not HAS_FEISHU:
        return jsonify({'ok': False, 'error': '飞书模块未安装', 'configured': False, 'cli_available': False}), 503
    bt_cfg = SiteConfig.query.filter_by(key='feishu_base_token').first()
    bt = bt_cfg.value if bt_cfg else None
    try:
        client = FeishuBaseClient(base_token=bt)
        cli_ok = client.check_cli()
        tables = []
        if bt:
            try:
                tables = client.list_tables()
            except Exception:
                tables = []
        return jsonify({
            'ok': True,
            'cli_available': cli_ok,
            'configured': bool(bt),
            'has_token': bool(bt),
            'tables': [{'id': t.get('id') or t.get('table_id'), 'name': t.get('name')} for t in tables] if tables else []
        })
    except FeishuCliError as e:
        return jsonify({'ok': False, 'cli_available': False, 'configured': bool(bt), 'error': str(e)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/feishu/config', methods=['GET'])
@login_required
def api_feishu_get_config():
    """获取飞书配置（token脱敏）"""
    if not HAS_FEISHU:
        return jsonify({'ok': False, 'error': '飞书模块未安装'}), 503
    configs = SiteConfig.query.filter(SiteConfig.key.like('feishu_%')).all()
    result = {}
    for c in configs:
        key = c.key
        if key == 'feishu_base_token' and c.value:
            v = c.value
            result[key] = v[:4] + '****' + v[-4:] if len(v) > 12 else '****'
        elif key in ('feishu_last_sync', 'feishu_sync_log'):
            try:
                result[key] = json.loads(c.value) if c.value else ([] if key == 'feishu_sync_log' else {})
            except (json.JSONDecodeError, TypeError):
                result[key] = c.value
        else:
            result[key] = c.value
    return jsonify({'ok': True, 'config': result})


@app.route('/api/feishu/config', methods=['POST'])
@login_required
def api_feishu_save_config():
    """保存飞书配置（支持 base_url 自动解析）"""
    if not HAS_FEISHU:
        return jsonify({'error': '飞书模块未安装'}), 503
    data = request.get_json()
    if not data:
        return jsonify({'error': '请提供配置数据'}), 400

    # 如果提供了 base_url，解析获取 base_token
    if data.get('base_url'):
        try:
            client = FeishuBaseClient()
            bt = client.resolve_base(data['base_url'])
            cfg = SiteConfig.query.filter_by(key='feishu_base_token').first()
            if cfg:
                cfg.value = bt
            else:
                db.session.add(SiteConfig(key='feishu_base_token', value=bt))
            log('飞书', '配置Base', f'通过URL解析: {bt[:8]}...')
        except (FeishuCliError, FeishuApiError) as e:
            return jsonify({'error': f'解析失败: {e}'}), 400

    # 如果直接提供了 base_token
    if data.get('base_token'):
        cfg = SiteConfig.query.filter_by(key='feishu_base_token').first()
        if cfg:
            cfg.value = data['base_token']
        else:
            db.session.add(SiteConfig(key='feishu_base_token', value=data['base_token']))
        log('飞书', '配置Base', '手动设置Token')

    # 保存 table_ids
    if data.get('table_ids'):
        for table_key, table_id in data['table_ids'].items():
            cfg = SiteConfig.query.filter_by(key=f'feishu_table_{table_key}').first()
            if cfg:
                cfg.value = str(table_id)
            else:
                db.session.add(SiteConfig(key=f'feishu_table_{table_key}', value=str(table_id)))

    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/feishu/init-tables', methods=['POST'])
@login_required
def api_feishu_init_tables():
    """在飞书中创建/确认标准表格"""
    if not HAS_FEISHU:
        return jsonify({'error': '飞书模块未安装'}), 503
    bt_cfg = SiteConfig.query.filter_by(key='feishu_base_token').first()
    if not bt_cfg or not bt_cfg.value:
        return jsonify({'error': '请先配置飞书Base连接'}), 400

    client = FeishuBaseClient(base_token=bt_cfg.value)
    try:
        table_map = client.ensure_tables()
        # 保存 table IDs 到配置
        for key, tid in table_map.items():
            cfg = SiteConfig.query.filter_by(key=f'feishu_table_{key}').first()
            if cfg:
                cfg.value = str(tid)
            else:
                db.session.add(SiteConfig(key=f'feishu_table_{key}', value=str(tid)))
        db.session.commit()
        log('飞书', '初始化表格', json.dumps(table_map))
        return jsonify({'ok': True, 'tables': table_map})
    except (FeishuCliError, FeishuApiError) as e:
        return jsonify({'error': f'初始化失败: {e}'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/feishu/sync', methods=['POST'])
@login_required
def api_feishu_sync():
    """同步数据到飞书多维表格"""
    if not HAS_FEISHU:
        return jsonify({'error': '飞书模块未安装'}), 503
    bt_cfg = SiteConfig.query.filter_by(key='feishu_base_token').first()
    if not bt_cfg or not bt_cfg.value:
        return jsonify({'error': '请先配置飞书Base连接'}), 400

    target = request.json.get('table', 'all')
    client = FeishuBaseClient(base_token=bt_cfg.value)
    results = {}
    sync_log_entries = []

    try:
        if target in ('activities', 'all'):
            tid_cfg = SiteConfig.query.filter_by(key='feishu_table_activities').first()
            if tid_cfg and tid_cfg.value:
                activities = [a.to_dict() for a in Activity.query.filter(Activity.deleted_at.is_(None)).all()]
                if activities:
                    r = client.sync_activities(tid_cfg.value, activities)
                else:
                    r = {'ok': True, 'total': 0, 'batches': 0}
                results['activities'] = r
                sync_log_entries.append({'table': '活动管理', 'status': 'success' if r['ok'] else 'error', 'count': r['total']})
            else:
                results['activities'] = {'ok': False, 'error': '未配置活动表ID', 'total': 0, 'batches': 0}

        if target in ('registrations', 'all'):
            tid_cfg = SiteConfig.query.filter_by(key='feishu_table_registrations').first()
            if tid_cfg and tid_cfg.value:
                registrations = [r.to_dict() for r in Registration.query.all()]
                if registrations:
                    r = client.sync_registrations(tid_cfg.value, registrations)
                else:
                    r = {'ok': True, 'total': 0, 'batches': 0}
                results['registrations'] = r
                sync_log_entries.append({'table': '报名记录', 'status': 'success' if r['ok'] else 'error', 'count': r['total']})
            else:
                results['registrations'] = {'ok': False, 'error': '未配置报名表ID', 'total': 0, 'batches': 0}

        if target in ('customers', 'all'):
            tid_cfg = SiteConfig.query.filter_by(key='feishu_table_customers').first()
            if tid_cfg and tid_cfg.value:
                customers = [c.to_dict() for c in Customer.query.all()]
                if customers:
                    r = client.sync_customers(tid_cfg.value, customers)
                else:
                    r = {'ok': True, 'total': 0, 'batches': 0}
                results['customers'] = r
                sync_log_entries.append({'table': '客户信息', 'status': 'success' if r['ok'] else 'error', 'count': r['total']})
            else:
                results['customers'] = {'ok': False, 'error': '未配置客户表ID', 'total': 0, 'batches': 0}

        # 保存同步日志
        now_str = datetime.datetime.now().strftime('%H:%M')
        for entry in sync_log_entries:
            entry['time'] = now_str
        log_cfg = SiteConfig.query.filter_by(key='feishu_sync_log').first()
        if log_cfg:
            try:
                existing = json.loads(log_cfg.value) if log_cfg.value else []
            except (json.JSONDecodeError, TypeError):
                existing = []
            existing.extend(sync_log_entries)
            log_cfg.value = json.dumps(existing[-50:], ensure_ascii=False)
        else:
            db.session.add(SiteConfig(key='feishu_sync_log', value=json.dumps(sync_log_entries, ensure_ascii=False)))

        db.session.commit()
        log('飞书', f'同步数据: {target}', json.dumps(results))
        return jsonify({'ok': True, 'results': results})

    except (FeishuCliError, FeishuApiError) as e:
        return jsonify({'error': f'同步失败: {e}'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/feishu/clear-logs', methods=['POST'])
@login_required
def api_feishu_clear_logs():
    """清空同步日志"""
    if not HAS_FEISHU:
        return jsonify({'ok': False, 'error': '飞书模块未安装'}), 503
    c = SiteConfig.query.filter_by(key='feishu_sync_log').first()
    if c:
        c.value = '[]'
        db.session.commit()
    return jsonify({'ok': True})
if __name__ == '__main__':
    with app.app_context():
        init_db()
        # 创建占位二维码（1x1透明像素，让img标签不报错）
        if not os.path.exists(QRCODE_PATH):
            from PIL import Image
            try:
                img = Image.new('RGBA', (300, 300), (245, 245, 245, 255))
                from PIL import ImageDraw, ImageFont
                draw = ImageDraw.Draw(img)
                draw.text((80, 130), "请上传二维码", fill=(150, 150, 150))
                img.save(QRCODE_PATH)
            except:
                pass
    app.run(debug=True, host='0.0.0.0', port=5000)
